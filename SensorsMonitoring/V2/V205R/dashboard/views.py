from django.shortcuts import render
from save_logs.models import *
from django.views.decorators.csrf import csrf_exempt
from django.http import StreamingHttpResponse,JsonResponse,HttpResponse
import time, json, jdatetime, datetime, platform, locale, os
from django.db.models import Q
from django.core.paginator import Paginator
from django.conf import settings
import csv
from io import BytesIO
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

def dashboard(request):
    all_devices = Device.objects.all()
    data = []
    for device in all_devices:
        response = {'device':device}
        sensors = Device_Sensor.objects.filter(device=device)
        response["sensors"] = sensors
        data.append(response)
    #print(data)
    context = {
        'data': data,
        'all_logs': len(data),
        'now': time.time() - 120,
    }
    return render(request, 'dashboard/dashboard.html', context)

def device_info_json(request,sensor_id):
    # sensor = Device_Sensor.objects.get(id=sensor_id)
    # data = SensorLogs.objects.filter(sensor=sensor)
    sensor = Device_Sensor.objects.get(id=sensor_id)
    data = SensorLogs.objects.filter(sensor=sensor)

    now = time.time()
    offset = 60*60
    chart_data = []
    #average_data = []
    for x in data:
        index = 0
        if not x.sensor.sensor_type == "status":
            for key, value in json.loads(x.data.replace("'",'"')).items():
                if not key == "timestamp" and not key == "type":
                    if key not in [x["type"] for x in chart_data]:
                        #average_data.append({"type":key,"average":value,"index":1})
                        chart_data.append({"type":key,"data":[],"timestamps":[],"time":float(x.LastUpdate)})
                    else:
                        if offset:
                            if chart_data[index]["time"] + offset < float(x.LastUpdate): 
                            #avg = (average_data[index]["average"] / average_data[index]["index"])
                                chart_data[index]["data"].append(float(f'{value:.2f}'))
                                chart_data[index]["timestamps"].append(float(x.CreationDateTime))
                                chart_data[index]["time"] = float(x.LastUpdate)
                        else:
                            chart_data[index]["data"].append(float(f'{value:.2f}'))
                            chart_data[index]["timestamps"].append(float(x.CreationDateTime))

                        # average_data[index]["average"] = valuez
                        # average_data[index]["index"] = 1
                            # print(average_data[index]["type"],value,average_data[index]["average"],average_data[index]["index"],average_data[index]["average"] / average_data[index]["index"])
                        # else:
                        #     # print(average_data[index]["type"],value,average_data[index]["average"],average_data[index]["index"],average_data[index]["average"] / average_data[index]["index"])
                        #     # open("average_data.txt","a").write(f"{average_data[index]['type']} {value} {average_data[index]['average']} {average_data[index]['index']} {average_data[index]['average'] / average_data[index]['index']}\n")
                        #     average_data[index]["average"] += value
                        #     average_data[index]["index"] += 1
                    index += 1
                #time.sleep(0.01)
    chart_data = [x for x in chart_data if x["data"]]
    return JsonResponse(list(chart_data),safe=False)
    return JsonResponse(list(data.values()),safe=False)

def device_info(request,device_id,sensor_id):
    device = Device.objects.get(id=device_id)
    sensor = Device_Sensor.objects.get(id=sensor_id)
    data = SensorLogs.objects.filter(sensor__device=device).filter(sensor=sensor)

    paginator = Paginator(data.order_by('-CreationDateTime'), 50)
    page_number = request.GET.get('page') or 1
    page_obj = paginator.get_page(page_number)
    
    # Load prediction data from JSON file
    prediction_data = load_prediction_data()
    
    context = {
        'data': page_obj,
        'device': device,
        'sensor': sensor,
        'all_logs': len(data),
        'page_obj': page_obj,
        'chart_data': [],
        'prediction_data': prediction_data.get('prediction_data', []),
        'pred_start_timestamps': prediction_data.get('pred_start_timestamps', 0),
    }
    return render(request, 'dashboard/device_info.html', context)

# Chart info JSON "device_info.html"
@csrf_exempt
def chart_info_json(request,sensor_id):
    if request.method == 'POST':
        sensor = Device_Sensor.objects.get(id=sensor_id)
        data = SensorLogs.objects.filter(sensor=sensor)
        filter_by_week = True if "week" in request.GET else False
        filter_by_month = True if "month" in request.GET else False
        filter_by_daily = True if "daily" in request.GET else False
        
        # Parse JSON data from request body
        try:
            request_data = json.loads(request.body.decode('utf-8'))
            filter_data = request_data.get('filter_data')
        except (json.JSONDecodeError, UnicodeDecodeError):
            # Fallback to POST data for backwards compatibility
            filter_data = request.POST.get('filter_data')
        target_filter_data = []
        if filter_by_week:
            for item in filter_data:
                if item['range'] == 'week':
                    min_value = item['min_value']
                    max_value = item['max_value']
                    filter_type = item['type']
                    target_filter_data.append(item)
        if filter_by_month:
            for item in filter_data:
                if item['range'] == 'month':
                    min_value = item['min_value']
                    max_value = item['max_value']
                    filter_type = item['type']
                    target_filter_data.append(item)
        if filter_by_daily:
            for item in filter_data:
                if item['range'] == 'daily':
                    min_value = item['min_value']
                    max_value = item['max_value']
                    filter_type = item['type']
                    target_filter_data.append(item)
        if not filter_by_week and not filter_by_month and not filter_by_daily:
            for item in filter_data:
                if item['range'] == 'hourly':
                    min_value = item['min_value']
                    max_value = item['max_value']
                    filter_type = item['type']
                    target_filter_data.append(item)

        now = time.time()
        filter_time = 60*60*24*7 if filter_by_week else 60*60*24*30 if filter_by_month else 60*60*24 if filter_by_daily else 60*60
        offset = False if filter_time == 3600 else 60*60
        
        chart_data = []
        min_set = False
        max_set = False
        for x in data.filter(CreationDateTime__gte=now - filter_time):
            index = 0
            if not x.sensor.sensor_type == "status":
                for key, value in json.loads(x.data.replace("'",'"')).items():
                    if not key == "timestamp" and not key == "type":
                        if key not in [x["type"] for x in chart_data]:
                            chart_data.append({"type":key,"data":[],"timestamps":[],"time":float(x.LastUpdate),"s":[]})
                        else:
                            min_value = False
                            max_value = False
                            if target_filter_data:
                                for item in target_filter_data:
                                    if item['type'] == key:
                                        min_value = item['min_value']
                                        max_value = item['max_value']
                                        break
                            if offset:
                                if chart_data[index]["time"] + offset < float(x.LastUpdate): 
                                    if min_value or max_value:
                                        if not min_value:
                                            min_value = float(value)
                                            min_set = True
                                        if not max_value:
                                            max_set = True
                                            max_value = float(value)
                                        if float(value) >= float(min_value) and float(value) <= float(max_value):
                                            chart_data[index]["data"].append(float(f'{value:.2f}'))
                                            chart_data[index]["timestamps"].append(str(jdatetime.datetime.fromgregorian(datetime=datetime.datetime.fromtimestamp(float(x.CreationDateTime))).strftime(" %m-%d %H:%M")))
                                            chart_data[index]["time"] = float(x.LastUpdate)
                                            chart_data[index]["s"].append(float(x.CreationDateTime))
                                    else:
                                        chart_data[index]["data"].append(float(f'{value:.2f}'))
                                        chart_data[index]["timestamps"].append(str(jdatetime.datetime.fromgregorian(datetime=datetime.datetime.fromtimestamp(float(x.CreationDateTime))).strftime(" %m-%d %H:%M")))
                                        chart_data[index]["time"] = float(x.LastUpdate)
                                        chart_data[index]["s"].append(float(x.CreationDateTime))
                            else:
                                if min_value or max_value:
                                    if not min_value:
                                        min_value = float(value)
                                        min_set = True
                                    if not max_value:
                                        max_set = True
                                        max_value = float(value)
                                    if float(value) >= float(min_value) and float(value) <= float(max_value):
                                        chart_data[index]["data"].append(float(f'{value:.2f}'))
                                        chart_data[index]["timestamps"].append(str(jdatetime.datetime.fromgregorian(datetime=datetime.datetime.fromtimestamp(float(x.CreationDateTime))).strftime(" %d %b %H:%M:%S")))
                                        chart_data[index]["s"].append(float(x.CreationDateTime))
                                else:
                                    chart_data[index]["data"].append(float(f'{value:.2f}'))
                                    chart_data[index]["timestamps"].append(str(jdatetime.datetime.fromgregorian(datetime=datetime.datetime.fromtimestamp(float(x.CreationDateTime))).strftime(" %d %b %H:%M:%S")))
                                    chart_data[index]["s"].append(float(x.CreationDateTime))
                            if min_value or max_value:
                                if min_set:
                                    min_value = False
                                    min_set = False
                                if max_set:
                                    max_value = False
                                    max_set = False
                        index += 1
        chart_data = [x for x in chart_data if x["data"]]
        for x in chart_data:
            if target_filter_data:
                for item in target_filter_data:
                    if x["type"] == item["type"]:
                        x["target_filter_data"] = item
                        break
            else:
                x["target_filter_data"] = []
        return JsonResponse(list(chart_data),safe=False)
    return JsonResponse({'status': 'error'})

@csrf_exempt
def ask_for_device(request):
    def event_stream():
        LAST_DATA_TIME = float(SensorLogs.objects.all().last().CreationDateTime)
        while True:
            time.sleep(1)
            target = SensorLogs.objects.all().last()
            if target:
                if float(target.CreationDateTime) > LAST_DATA_TIME:
                    print("some data recived")
                    resp = {
                        "device_id": target.sensor.device.device_id,
                        "ip_address": target.sensor.device.ip_address,
                        "Is_Known": target.sensor.device.Is_Known,
                        "sensor_type": target.sensor.sensor_type,
                        "location": target.sensor.device.location,
                        "name": target.sensor.device.name,
                        "data": target.data,
                        "CreationDateTime": str(jdatetime.datetime.fromgregorian(datetime=datetime.datetime.fromtimestamp(float(target.CreationDateTime))).strftime("%a, %d %b %Y %H:%M:%S")),
                    }
                    LAST_DATA_TIME = float(target.CreationDateTime)
                    yield 'data: %s\n\n' % json.dumps(resp)
                yield 'data: {}\n\n'
            else:
                # Send keep-alive to maintain connection
                yield 'data: {}\n\n'
            time.sleep(1)

    return StreamingHttpResponse(event_stream(), content_type='text/event-stream')

@csrf_exempt
def add_device(request):
    if request.method == 'POST':
        device_id = request.POST.get('device_id')
        name = request.POST.get('name')
        location = request.POST.get('location')
        print(device_id,"__________")
        if name or location:
            target = Device.objects.filter(device_id=device_id).last()
            target.name = name
            target.location = location
            target.Is_Known = True
            target.save()
            print("device added successfully | now we now the device name and location")
            return JsonResponse({'status': 'ok'})
        else:
            return JsonResponse({'status': 'error'})
    return JsonResponse({'status': 'error'})

@csrf_exempt
def get_sensor_status(request):
    """API endpoint to get sensor status data for real-time updates"""
    try:
        sensors_status = []
        current_time = time.time()
        two_minutes_ago = current_time - 120
        
        # Get all sensors with their latest log
        sensors = Device_Sensor.objects.all()
        
        for sensor in sensors:
            if sensor.sensor_type != "status":
                last_log = sensor.sensor_logs.last()
                if last_log:
                    is_online = last_log.CreationDateTime > two_minutes_ago
                    sensors_status.append({
                        'sensor_id': sensor.id,
                        'last_update': last_log.CreationDateTime,
                        'is_online': is_online
                    })
                else:
                    sensors_status.append({
                        'sensor_id': sensor.id,
                        'last_update': None,
                        'is_online': False
                    })
        
        return JsonResponse({
            'status': 'success',
            'current_time': current_time,
            'sensors': sensors_status
        })
    
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })

def load_prediction_data():
    """Load prediction data from JSON file"""
    prediction_file = os.path.join(settings.BASE_DIR, 'prediction.json')
    try:
        with open(prediction_file, 'r') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {
            "prediction_data": [],
            "pred_start_timestamps": 0
        }

def save_prediction_data(data):
    """Save prediction data to JSON file"""
    prediction_file = os.path.join(settings.BASE_DIR, 'prediction.json')
    with open(prediction_file, 'w') as f:
        json.dump(data, f, indent=4)

@csrf_exempt
def update_prediction(request):
    """View to update prediction data"""
    if request.method == 'GET':
        # Load current data for display
        current_data = load_prediction_data()
        context = {
            'prediction_data': current_data.get('prediction_data', []),
            'pred_start_timestamps': current_data.get('pred_start_timestamps', 0)
        }
        return render(request, 'dashboard/update_prediction.html', context)
    
    elif request.method == 'POST':
        try:
            # Get data from form
            prediction_data_str = request.POST.get('prediction_data', '').strip()
            pred_start_timestamps = request.POST.get('pred_start_timestamps', '0').strip()
            
            # Parse prediction data (list of floats)
            if prediction_data_str:
                # Remove brackets and split by comma
                prediction_data_str = prediction_data_str.strip('[]')
                prediction_data = [float(x.strip()) for x in prediction_data_str.split(',') if x.strip()]
            else:
                prediction_data = []
            
            # Parse timestamp
            pred_start_timestamps = float(pred_start_timestamps) if pred_start_timestamps else 0
            
            # Save to JSON file
            data = {
                "prediction_data": prediction_data,
                "pred_start_timestamps": pred_start_timestamps
            }
            save_prediction_data(data)
            
            return JsonResponse({'status': 'success', 'message': 'Prediction data updated successfully'})
            
        except ValueError as e:
            return JsonResponse({'status': 'error', 'message': f'Invalid data format: {str(e)}'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Error saving data: {str(e)}'})

@csrf_exempt
def get_device_data(request, device_id):
    """Get device data for editing"""
    try:
        device = Device.objects.get(id=device_id)
        return JsonResponse({
            'status': 'success',
            'device': {
                'id': device.id,
                'device_id': device.device_id,
                'name': device.name,
                'location': device.location,
                'ip_address': device.ip_address,
                'description': device.description,
            }
        })
    except Device.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Device not found'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@csrf_exempt
def get_sensor_data(request, sensor_id):
    """Get sensor data for editing"""
    try:
        sensor = Device_Sensor.objects.get(id=sensor_id)
        return JsonResponse({
            'status': 'success',
            'sensor': {
                'id': sensor.id,
                'sensor_type': sensor.sensor_type,
                'description': sensor.description,
            }
        })
    except Device_Sensor.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Sensor not found'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@csrf_exempt
def update_device(request):
    """Update device information"""
    if request.method == 'POST':
        try:
            device_id = request.POST.get('device_id')
            if not device_id:
                return JsonResponse({'status': 'error', 'message': 'Device ID is required'})
            
            device = Device.objects.get(id=device_id)
            
            # Update fields if provided
            if 'name' in request.POST:
                device.name = request.POST.get('name', '').strip() or None
            if 'location' in request.POST:
                device.location = request.POST.get('location', '').strip() or None
            if 'ip_address' in request.POST:
                device.ip_address = request.POST.get('ip_address', '').strip() or None
            if 'description' in request.POST:
                device.description = request.POST.get('description', '').strip() or None
            
            device.save()
            
            return JsonResponse({'status': 'success', 'message': 'Device updated successfully'})
        except Device.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Device not found'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Only POST method allowed'})

@csrf_exempt
def update_sensor(request):
    """Update sensor information"""
    if request.method == 'POST':
        try:
            sensor_id = request.POST.get('sensor_id')
            if not sensor_id:
                return JsonResponse({'status': 'error', 'message': 'Sensor ID is required'})
            
            sensor = Device_Sensor.objects.get(id=sensor_id)
            
            # Update fields if provided
            if 'sensor_type' in request.POST:
                sensor_type = request.POST.get('sensor_type', '').strip()
                if sensor_type:
                    sensor.sensor_type = sensor_type
                else:
                    return JsonResponse({'status': 'error', 'message': 'Sensor type is required'})
            
            if 'description' in request.POST:
                sensor.description = request.POST.get('description', '').strip() or None
            
            sensor.save()
            
            return JsonResponse({'status': 'success', 'message': 'Sensor updated successfully'})
        except Device_Sensor.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Sensor not found'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Only POST method allowed'})


@csrf_exempt
def export_data(request, sensor_id):
    """Export chart data as Excel file with separate sheets for each data type"""
    try:
        sensor = Device_Sensor.objects.get(id=sensor_id)
        
        # Get filter parameters from request
        filter_by_week = True if "week" in request.GET else False
        filter_by_month = True if "month" in request.GET else False
        filter_by_daily = True if "daily" in request.GET else False
        
        # Parse JSON data from request body
        try:
            request_data = json.loads(request.body.decode('utf-8'))
            filter_data = request_data.get('filter_data', [])
            export_choices = request_data.get('export_choices', '1')
        except (json.JSONDecodeError, UnicodeDecodeError):
            filter_data = []
            export_choices = request.POST.get('export_choices', '1')
        # Determine target filter data based on time range
        target_filter_data = []
        time_range_name = "Hourly"
        if filter_by_week:
            time_range_name = "Weekly"
            for item in filter_data:
                if item['range'] == 'week':
                    target_filter_data.append(item)
        elif filter_by_month:
            time_range_name = "Monthly"
            for item in filter_data:
                if item['range'] == 'month':
                    target_filter_data.append(item)
        elif filter_by_daily:
            time_range_name = "Daily"
            for item in filter_data:
                if item['range'] == 'daily':
                    target_filter_data.append(item)
        else:
            time_range_name = "Hourly"
            for item in filter_data:
                if item['range'] == 'hourly':
                    target_filter_data.append(item)
        
        # Get sensor data
        data = SensorLogs.objects.filter(sensor=sensor)
        
        # Calculate time filter
        now = time.time()
        filter_time = 60*60*24*7 if filter_by_week else 60*60*24*30 if filter_by_month else 60*60*24 if filter_by_daily else 60*60
        offset = False if filter_time == 3600 else 60*60
        print(export_choices,"export_choices")
        if not export_choices == "1":
            offset = False
        # Build chart data (same logic as chart_info_json)
        chart_data = []
        min_set = False
        max_set = False
        
        for x in data.filter(CreationDateTime__gte=now - filter_time):
            index = 0
            if not x.sensor.sensor_type == "status":
                for key, value in json.loads(x.data.replace("'",'"')).items():
                    if not key == "timestamp" and not key == "type":
                        if key not in [x["type"] for x in chart_data]:
                            chart_data.append({"type":key,"data":[],"timestamps":[],"time":float(x.LastUpdate),"s":[]})
                        else:
                            min_value = False
                            max_value = False
                            if target_filter_data:
                                for item in target_filter_data:
                                    if item['type'] == key:
                                        min_value = item['min_value']
                                        max_value = item['max_value']
                                        break
                            if offset:
                                if chart_data[index]["time"] + offset < float(x.LastUpdate): 
                                    if min_value or max_value:
                                        if not min_value:
                                            min_value = float(value)
                                            min_set = True
                                        if not max_value:
                                            max_set = True
                                            max_value = float(value)
                                        if float(value) >= float(min_value) and float(value) <= float(max_value):
                                            chart_data[index]["data"].append(float(f'{value:.2f}'))
                                            chart_data[index]["timestamps"].append(str(jdatetime.datetime.fromgregorian(datetime=datetime.datetime.fromtimestamp(float(x.CreationDateTime))).strftime("%Y-%m-%d %H:%M:%S")))
                                            chart_data[index]["time"] = float(x.LastUpdate)
                                            chart_data[index]["s"].append(float(x.CreationDateTime))
                                    else:
                                        chart_data[index]["data"].append(float(f'{value:.2f}'))
                                        chart_data[index]["timestamps"].append(str(jdatetime.datetime.fromgregorian(datetime=datetime.datetime.fromtimestamp(float(x.CreationDateTime))).strftime("%Y-%m-%d %H:%M:%S")))
                                        chart_data[index]["time"] = float(x.LastUpdate)
                                        chart_data[index]["s"].append(float(x.CreationDateTime))
                            else:
                                if min_value or max_value:
                                    if not min_value:
                                        min_value = float(value)
                                        min_set = True
                                    if not max_value:
                                        max_set = True
                                        max_value = float(value)
                                    if float(value) >= float(min_value) and float(value) <= float(max_value):
                                        chart_data[index]["data"].append(float(f'{value:.2f}'))
                                        chart_data[index]["timestamps"].append(str(jdatetime.datetime.fromgregorian(datetime=datetime.datetime.fromtimestamp(float(x.CreationDateTime))).strftime("%Y-%m-%d %H:%M:%S")))
                                        chart_data[index]["s"].append(float(x.CreationDateTime))
                                else:
                                    chart_data[index]["data"].append(float(f'{value:.2f}'))
                                    chart_data[index]["timestamps"].append(str(jdatetime.datetime.fromgregorian(datetime=datetime.datetime.fromtimestamp(float(x.CreationDateTime))).strftime("%Y-%m-%d %H:%M:%S")))
                                    chart_data[index]["s"].append(float(x.CreationDateTime))
                            if min_value or max_value:
                                if min_set:
                                    min_value = False
                                    min_set = False
                                if max_set:
                                    max_value = False
                                    max_set = False
                        index += 1
        
        chart_data = [x for x in chart_data if x["data"]]
        if not chart_data:
            return JsonResponse({'status': 'error', 'message': 'دیتا وجود ندارد'})
        
        # Check if openpyxl is available for Excel export
        if OPENPYXL_AVAILABLE:
            # Create Excel file with multiple sheets
            output = BytesIO()
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Create a sheet for each data type
            for data_type_info in chart_data:
                data_type = data_type_info['type']
                values = data_type_info['data']
                timestamps = data_type_info['timestamps']
                
                # Create sheet for this data type
                ws = wb.create_sheet(title=data_type[:31])  # Excel sheet name limit is 31 chars
                
                # Style header row
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # Add headers
                ws['A1'] = 'Timestamp'
                ws['B1'] = f'{data_type} Value'
                ws['A1'].fill = header_fill
                ws['B1'].fill = header_fill
                ws['A1'].font = header_font
                ws['B1'].font = header_font
                ws['A1'].alignment = header_alignment
                ws['B1'].alignment = header_alignment
                
                # Add data rows
                for i, (timestamp, value) in enumerate(zip(timestamps, values), start=2):
                    ws[f'A{i}'] = timestamp
                    ws[f'B{i}'] = value
                
                # Adjust column widths
                ws.column_dimensions['A'].width = 25
                ws.column_dimensions['B'].width = 15
            
            # Save workbook to BytesIO
            wb.save(output)
            output.seek(0)
            
            # Create response
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # Generate filename with sensor info and time range
            filename = f'{sensor.sensor_type}_{time_range_name}_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
        else:
            # Fallback to CSV if openpyxl is not available
            # Export only the first data type or combine all
            response = HttpResponse(content_type='text/csv')
            filename = f'{sensor.sensor_type}_{time_range_name}_{jdatetime.datetime.fromgregorian(datetime=datetime.datetime.now()).strftime("%Y%m%d_%H%M%S")}.csv'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            writer = csv.writer(response)
            
            # Write data for each type
            for data_type_info in chart_data:
                data_type = data_type_info['type']
                values = data_type_info['data']
                timestamps = data_type_info['timestamps']
                
                writer.writerow([])  # Empty row separator
                writer.writerow([f'=== {data_type} ==='])
                writer.writerow(['Timestamp', f'{data_type} Value'])
                
                for timestamp, value in zip(timestamps, values):
                    writer.writerow([timestamp, value])
            
            return response
            
    except Device_Sensor.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'سنسور یافت نشد'})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'status': 'error', 'message': str(e)})