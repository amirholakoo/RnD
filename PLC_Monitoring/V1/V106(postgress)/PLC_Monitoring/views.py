from django.http import JsonResponse, StreamingHttpResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from django.shortcuts import render
from .models import *
from django.utils import timezone
import json
import time
import jdatetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def dashboard(request):
    context = {
        'all_plcs': PLC.objects.all(),
    }
    return render(request, 'plc_monitoring/dashboard.html', context)

def get_plcs(request):
    """Get all PLCs"""
    plcs = PLC.objects.all()
    data = []
    for plc in plcs:
        data.append({
            'id': plc.id,
            'device_id': plc.device_id,
            'name': plc.name,
            'ip_address': plc.ip_address,
            'location': plc.location,
            'Is_Known': plc.Is_Known,
            'description': plc.description,
            'CreationDateTime': plc.CreationDateTime,
            'LastUpdate': plc.LastUpdate
        })
    return JsonResponse({'status': 'ok', 'data': data})

@csrf_exempt
def create_plc(request):
    """Create a new PLC"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    try:
        device_id = request.POST.get('device_id')
        name = request.POST.get('name', '')
        ip_address = request.POST.get('ip_address', '')
        location = request.POST.get('location', '')
        description = request.POST.get('description', '')
        is_known = request.POST.get('Is_Known') == 'on'
        
        if not device_id:
            return JsonResponse({'status': 'error', 'message': 'شناسه دستگاه الزامی است'})
        
        if PLC.objects.filter(device_id=device_id).exists():
            return JsonResponse({'status': 'error', 'message': 'شناسه دستگاه تکراری است'})
        
        plc = PLC.objects.create(
            device_id=device_id,
            name=name,
            ip_address=ip_address,
            location=location,
            description=description,
            Is_Known=is_known
        )
        
        return JsonResponse({
            'status': 'ok',
            'data': {
                'id': plc.id,
                'device_id': plc.device_id,
                'name': plc.name,
                'ip_address': plc.ip_address,
                'location': plc.location,
                'Is_Known': plc.Is_Known
            }
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@csrf_exempt
def update_plc(request):
    """Update an existing PLC"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    try:
        plc_id = request.POST.get('id')
        device_id = request.POST.get('device_id')
        name = request.POST.get('name', '')
        ip_address = request.POST.get('ip_address', '')
        location = request.POST.get('location', '')
        description = request.POST.get('description', '')
        is_known = request.POST.get('Is_Known') == 'on'
        
        if not plc_id or not device_id:
            return JsonResponse({'status': 'error', 'message': 'فیلدهای اجباری را پر کنید'})
        
        plc = PLC.objects.get(id=plc_id)
        
        if PLC.objects.filter(device_id=device_id).exclude(id=plc_id).exists():
            return JsonResponse({'status': 'error', 'message': 'شناسه دستگاه تکراری است'})
        
        plc.device_id = device_id
        plc.name = name
        plc.ip_address = ip_address
        plc.location = location
        plc.description = description
        plc.Is_Known = is_known
        plc.save()
        
        return JsonResponse({
            'status': 'ok',
            'data': {
                'id': plc.id,
                'device_id': plc.device_id,
                'name': plc.name,
                'ip_address': plc.ip_address,
                'location': plc.location,
                'Is_Known': plc.Is_Known
            }
        })
    except PLC.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'دستگاه یافت نشد'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@csrf_exempt
def delete_plc(request):
    """Delete a PLC"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    try:
        plc_id = request.POST.get('id')
        
        if not plc_id:
            return JsonResponse({'status': 'error', 'message': 'شناسه دستگاه الزامی است'})
        
        plc = PLC.objects.get(id=plc_id)
        plc.delete()
        
        return JsonResponse({'status': 'ok', 'message': 'دستگاه با موفقیت حذف شد'})
    except PLC.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'دستگاه یافت نشد'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

def get_plc_keys(request):
    """Get all PLC keys"""
    keys = PLC_Keys.objects.all()
    
    data = []
    for key in keys:
        data.append({
            'id': key.id,
            'name': key.name,
            'fa_name': key.fa_name,
            'key': key.key,
            'value': key.value,
            'description': key.description,
            'CreationDateTime': key.CreationDateTime,
            'LastUpdate': key.LastUpdate
        })
    
    return JsonResponse({'status': 'ok', 'data': data})

@csrf_exempt
def create_plc_key(request):
    """Create a new PLC key"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    try:
        name = request.POST.get('name', '')
        fa_name = request.POST.get('fa_name', '')
        key = request.POST.get('key')
        value = request.POST.get('value', '')
        description = request.POST.get('description', '')
        
        if not key:
            return JsonResponse({'status': 'error', 'message': 'نام کلید الزامی است'})
        
        plc_key = PLC_Keys.objects.create(
            name=name,
            fa_name=fa_name,
            key=key,
            value=value,
            description=description
        )
        
        return JsonResponse({
            'status': 'ok',
            'data': {
                'id': plc_key.id,
                'name': plc_key.name,
                'fa_name': plc_key.fa_name,
                'key': plc_key.key,
                'value': plc_key.value,
                'description': plc_key.description
            }
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@csrf_exempt
def update_plc_key(request):
    """Update an existing PLC key"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    try:
        key_id = request.POST.get('id')
        name = request.POST.get('name', '')
        fa_name = request.POST.get('fa_name', '')
        key = request.POST.get('key')
        value = request.POST.get('value', '')
        description = request.POST.get('description', '')
        
        if not key_id or not key:
            return JsonResponse({'status': 'error', 'message': 'فیلدهای اجباری را پر کنید'})
        
        plc_key = PLC_Keys.objects.get(id=key_id)
        plc_key.name = name
        plc_key.fa_name = fa_name
        plc_key.key = key
        plc_key.value = value
        plc_key.description = description
        plc_key.save()
        
        return JsonResponse({
            'status': 'ok',
            'data': {
                'id': plc_key.id,
                'name': plc_key.name,
                'fa_name': plc_key.fa_name,
                'key': plc_key.key,
                'value': plc_key.value,
                'description': plc_key.description
            }
        })
    except PLC_Keys.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'کلید یافت نشد'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@csrf_exempt
def delete_plc_key(request):
    """Delete a PLC key"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    try:
        key_id = request.POST.get('id')
        
        if not key_id:
            return JsonResponse({'status': 'error', 'message': 'شناسه کلید الزامی است'})
        
        plc_key = PLC_Keys.objects.get(id=key_id)
        plc_key.delete()
        
        return JsonResponse({'status': 'ok', 'message': 'کلید با موفقیت حذف شد'})
    except PLC_Keys.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'کلید یافت نشد'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

def sse_plc_logs(request):
    """SSE endpoint for streaming PLC_Logs in real-time (new logs + updated logs)"""
    def event_stream():
        last_id = 0
        last_update_time = timezone.now()
        
        while True:
            logs_data = []
            
            # Get new logs
            new_logs = PLC_Logs.objects.filter(id__gt=last_id).select_related('plc').order_by('-id')[:50]
            if new_logs:
                last_id = new_logs[0].id
                for log in reversed(new_logs):
                    logs_data.append({
                        'id': log.id,
                        'plc_id': log.plc.id if log.plc else None,
                        'plc_device_id': log.plc.device_id if log.plc else 'نامشخص',
                        'plc_name': log.plc.name if log.plc else '',
                        'data': log.data,
                        'json_data': log.json_data,
                        'CreationDateTime': log.CreationDateTime,
                        'LastUpdate': log.LastUpdate,
                        'is_update': False
                    })
            
            # Get updated logs (LastUpdate changed since last check)
            updated_logs = PLC_Logs.objects.filter(
                id__lte=last_id,
                LastUpdate__gt=last_update_time
            ).select_related('plc').order_by('-LastUpdate')[:20]
            
            for log in updated_logs:
                logs_data.append({
                    'id': log.id,
                    'plc_id': log.plc.id if log.plc else None,
                    'plc_device_id': log.plc.device_id if log.plc else 'نامشخص',
                    'plc_name': log.plc.name if log.plc else '',
                    'data': log.data,
                    'json_data': log.json_data,
                    'CreationDateTime': log.CreationDateTime,
                    'LastUpdate': log.LastUpdate,
                    'is_update': True
                })
            
            last_update_time = timezone.now()
            
            if logs_data:
                yield f"data: {json.dumps(logs_data, ensure_ascii=False)}\n\n"
            
            time.sleep(1)
    
    response = StreamingHttpResponse(event_stream(), content_type='text/event-stream')
    response['Cache-Control'] = 'no-cache'
    response['X-Accel-Buffering'] = 'no'
    return response

def get_initial_logs(request):
    """Get initial PLC_Logs for table"""
    limit = int(request.GET.get('limit', 50))
    logs = PLC_Logs.objects.select_related('plc').order_by('-id')[:limit]
    logs_data = []
    for log in logs:
        logs_data.append({
            'id': log.id,
            'plc_id': log.plc.id if log.plc else None,
            'plc_device_id': log.plc.device_id if log.plc else 'نامشخص',
            'plc_name': log.plc.name if log.plc else '',
            'data': log.data,
            'json_data': log.json_data,
            'CreationDateTime': log.CreationDateTime,
            'LastUpdate': log.LastUpdate
        })
    return JsonResponse({'status': 'ok', 'data': logs_data})

def live_settings(request):
    """Live PLC settings page for a specific PLC with pagination and search"""
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    plc_id = request.GET.get('plc')
    if not plc_id:
        return render(request, 'plc_monitoring/live_settings.html', {'plc': None})
    
    try:
        plc = PLC.objects.get(id=plc_id)
        rolls = Rolls.objects.filter(plc=plc, Is_Deleted=False).order_by('-CreationDateTime')
        
        # Search by roll number
        search_query = request.GET.get('search', '').strip()
        if search_query:
            rolls = rolls.filter(roll_number__icontains=search_query)
        
        # Pagination - 50 per page
        paginator = Paginator(rolls, 50)
        page = request.GET.get('page', 1)
        try:
            rolls_page = paginator.page(page)
        except PageNotAnInteger:
            rolls_page = paginator.page(1)
        except EmptyPage:
            rolls_page = paginator.page(paginator.num_pages)
            
    except PLC.DoesNotExist:
        plc = None
        rolls_page = None
        search_query = ''
        paginator = None
    
    # Get key translations for charts
    key_translations = {}
    plc_keys = PLC_Keys.objects.all()
    for pk in plc_keys:
        key_translations[pk.key] = pk.fa_name or pk.name or pk.key
    
    # Get excluded keys for settings panel
    excluded_keys = set(ChartExcludedKeys.objects.values_list('key', flat=True))
    keys_with_status = []
    for pk in plc_keys:
        keys_with_status.append({
            'key': pk.key,
            'fa_name': pk.fa_name or pk.name or pk.key,
            'is_excluded': pk.key in excluded_keys
        })
    
    return render(request, 'plc_monitoring/live_settings.html', {
        'plc': plc, 
        'rolls': rolls_page,
        'search_query': search_query if plc else '',
        'paginator': paginator,
        'key_translations': json.dumps(key_translations, ensure_ascii=False),
        'keys_with_status': json.dumps(keys_with_status, ensure_ascii=False)
    })

def get_historical_chart_data(request):
    """Get historical chart data for a PLC based on time range"""
    plc_id = request.GET.get('plc')
    time_range = request.GET.get('range', '1h')
    interval = int(request.GET.get('interval', 60))  # Default 60 seconds
    
    if not plc_id:
        return JsonResponse({'status': 'error', 'message': 'PLC ID required'})
    
    # Calculate time threshold based on range
    now = timezone.now()
    time_thresholds = {
        '1h': now - 3600,
        '8h': now - 28800,
        '24h': now - 86400,
        'week': now - 604800,
        'month': now - 2592000,
        'year': now - 31536000
    }
    threshold = time_thresholds.get(time_range, now - 3600)
    
    # Get excluded keys
    excluded_keys = set(ChartExcludedKeys.objects.values_list('key', flat=True))
    
    # Get key translations
    key_translations = {}
    plc_keys = PLC_Keys.objects.all()
    for pk in plc_keys:
        key_translations[pk.key] = pk.fa_name or pk.name or pk.key
    
    try:
        # Get rolls in time range
        rolls = Rolls.objects.filter(
            plc_id=plc_id,
            Is_Deleted=False,
            CreationDateTime__gte=threshold
        ).order_by('CreationDateTime')
        
        # Get all logs for these rolls
        roll_ids = list(rolls.values_list('id', flat=True))
        logs = PLC_Logs.objects.filter(
            roll_id__in=roll_ids,
            is_running=True
        ).order_by('CreationDateTime')
        
        # Collect all unique keys
        all_keys = set()
        for log in logs:
            if log.json_data:
                for key in log.json_data.keys():
                    if key not in excluded_keys:
                        all_keys.add(key)
        
        # Build raw data grouped by interval (last value in each interval)
        # interval_data[key][interval_timestamp] = last_value
        interval_data = {key: {} for key in all_keys}
        
        for log in logs:
            if not log.CreationDateTime or not log.json_data:
                continue
            # Round timestamp to interval
            interval_ts = int(log.CreationDateTime.timestamp() // interval) * interval * 1000
            for key in all_keys:
                if key in log.json_data:
                    try:
                        value = float(log.json_data[key])
                        interval_data[key][interval_ts] = value  # Last value wins
                    except (ValueError, TypeError):
                        pass
        
        # Get order_index for keys (use 9999 as default for unordered keys)
        key_order = {}
        for pk in plc_keys:
            key_order[pk.key] = pk.order_index if pk.order_index is not None else 9999
        
        # Convert to series format
        chart_series = []
        for key, data in interval_data.items():
            if data:
                sorted_data = [{'x': ts, 'y': val} for ts, val in sorted(data.items())]
                fa_name = key_translations.get(key, key)
                order = key_order.get(key, 9999)
                chart_series.append({'name': key, 'fa_name': fa_name, 'data': sorted_data, 'order_index': order})
        
        # Sort by order_index, then by name for consistent ordering
        chart_series.sort(key=lambda x: (x['order_index'], x['name']))
        
        # Build roll annotations (first log of each roll)
        roll_annotations = []
        for roll in rolls:
            first_log = PLC_Logs.objects.filter(roll=roll, is_running=True).order_by('CreationDateTime').first()
            if first_log and first_log.CreationDateTime:
                roll_annotations.append({
                    'x': int(first_log.CreationDateTime.timestamp() * 1000),
                    'label': f'{roll.roll_number or roll.id}'
                })
        
        # Build roll breaks annotations
        break_annotations = []
        roll_breaks = Roll_Breaks.objects.filter(roll_id__in=roll_ids).order_by('CreationDateTime')
        for rb in roll_breaks:
            if rb.CreationDateTime:
                break_annotations.append({
                    'x': int(rb.CreationDateTime.timestamp() * 1000),
                    'label': rb.break_reason or 'پارگی'
                })
        
        # Build stopped ranges (is_running=False periods, only if > 5 minutes)
        stopped_ranges = []
        all_logs = PLC_Logs.objects.filter(
            roll_id__in=roll_ids
        ).order_by('CreationDateTime').values('CreationDateTime', 'is_running')
        
        range_start = None
        MIN_STOPPED_DURATION_MS = 300000  # 5 minutes in milliseconds
        
        for log in all_logs:
            if not log['CreationDateTime']:
                continue
            if not log['is_running'] and range_start is None:
                range_start = int(log['CreationDateTime'].timestamp() * 1000)
            elif log['is_running'] and range_start is not None:
                range_end = int(log['CreationDateTime'].timestamp() * 1000)
                if range_end - range_start >= MIN_STOPPED_DURATION_MS:
                    stopped_ranges.append({'x': range_start, 'x2': range_end})
                range_start = None
        # Handle case where last logs are is_running=False
        if range_start is not None:
            range_end = int(timezone.now().timestamp() * 1000)
            if range_end - range_start >= MIN_STOPPED_DURATION_MS:
                stopped_ranges.append({'x': range_start, 'x2': range_end})
        
        return JsonResponse({
            'status': 'ok',
            'chart_series': chart_series,
            'roll_annotations': roll_annotations,
            'break_annotations': break_annotations,
            'stopped_ranges': stopped_ranges
        })
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

def roll_detail(request, roll_id):
    """Roll detail page showing roll info and plc_setting"""
    # Get dynamic excluded keys from database
    excluded_keys = set(ChartExcludedKeys.objects.values_list('key', flat=True))
    interval = int(request.GET.get('interval', 60))  # Default 60 seconds
    
    try:
        roll = Rolls.objects.select_related('plc').get(id=roll_id)
        logs = PLC_Logs.objects.filter(roll=roll).filter(is_running=True).order_by('CreationDateTime')
        
        # Collect all unique keys from json_data
        all_keys = set()
        for log in logs:
            if log.json_data:
                for key in log.json_data.keys():
                    if key not in excluded_keys:
                        all_keys.add(key)
        
        # Get key translations from PLC_Keys
        key_translations = {}
        plc_keys = PLC_Keys.objects.filter(key__in=all_keys)
        for pk in plc_keys:
            key_translations[pk.key] = pk.fa_name or pk.name or pk.key
        
        # Build data grouped by interval (last value in each interval)
        interval_data = {key: {} for key in all_keys}
        
        for log in logs:
            if not log.CreationDateTime or not log.json_data:
                continue
            # Round timestamp to interval
            interval_ts = int(log.CreationDateTime.timestamp() // interval) * interval * 1000
            for key in all_keys:
                if key in log.json_data:
                    try:
                        value = float(log.json_data[key])
                        interval_data[key][interval_ts] = value  # Last value wins
                    except (ValueError, TypeError):
                        pass
        
        # Convert to list format for ApexCharts with fa_name
        chart_series = []
        for key, data in interval_data.items():
            if data:
                sorted_data = [{'x': ts, 'y': val} for ts, val in sorted(data.items())]
                fa_name = key_translations.get(key, key)
                chart_series.append({'name': key, 'fa_name': fa_name, 'data': sorted_data})
        
        # Build roll breaks annotations
        break_annotations = []
        roll_breaks = Roll_Breaks.objects.filter(roll=roll).order_by('CreationDateTime')
        for rb in roll_breaks:
            if rb.CreationDateTime:
                break_annotations.append({
                    'x': int(rb.CreationDateTime.timestamp() * 1000),
                    'label': rb.break_reason or 'پارگی'
                })
        
        # Build stopped ranges (is_running=False periods, only if > 5 minutes)
        stopped_ranges = []
        all_logs = PLC_Logs.objects.filter(roll=roll).order_by('CreationDateTime').values('CreationDateTime', 'is_running')
        
        range_start = None
        MIN_STOPPED_DURATION_MS = 300000  # 5 minutes in milliseconds
        
        for log in all_logs:
            if not log['CreationDateTime']:
                continue
            if not log['is_running'] and range_start is None:
                range_start = int(log['CreationDateTime'].timestamp() * 1000)
            elif log['is_running'] and range_start is not None:
                range_end = int(log['CreationDateTime'].timestamp() * 1000)
                if range_end - range_start >= MIN_STOPPED_DURATION_MS:
                    stopped_ranges.append({'x': range_start, 'x2': range_end})
                range_start = None
        if range_start is not None:
            range_end = int(timezone.now().timestamp() * 1000)
            if range_end - range_start >= MIN_STOPPED_DURATION_MS:
                stopped_ranges.append({'x': range_start, 'x2': range_end})
        
    except Rolls.DoesNotExist:
        roll = None
        chart_series = []
        break_annotations = []
        stopped_ranges = []
    
    # Get all PLC_Keys with excluded status for UI
    all_plc_keys = PLC_Keys.objects.all()
    keys_with_status = []
    for pk in all_plc_keys:
        keys_with_status.append({
            'key': pk.key,
            'fa_name': pk.fa_name or pk.name or pk.key,
            'is_excluded': pk.key in excluded_keys
        })
    
    return render(request, 'plc_monitoring/roll_detail.html', {
        'roll': roll,
        'chart_series': json.dumps(chart_series, ensure_ascii=False),
        'keys_with_status': json.dumps(keys_with_status, ensure_ascii=False),
        'stopped_ranges': json.dumps(stopped_ranges, ensure_ascii=False),
        'break_annotations': json.dumps(break_annotations, ensure_ascii=False),
        'current_interval': interval
    })

def get_plc_settings(request):
    """Get all PLCs with their settings"""
    plcs = PLC.objects.all()
    data = []
    for plc in plcs:
        data.append({
            'id': plc.id,
            'device_id': plc.device_id,
            'name': plc.name,
            'setting': plc.setting,
            'LastUpdate': plc.LastUpdate
        })
    return JsonResponse({'status': 'ok', 'data': data})

def sse_plc_settings(request):
    """SSE endpoint for streaming PLC settings in real-time"""
    plc_id = request.GET.get('plc')
    
    def event_stream():
        last_update_time = 0
        
        while True:
            try:
                plc = PLC.objects.get(id=plc_id)
                
                if plc.LastUpdate and plc.LastUpdate > last_update_time:
                    last_update_time = plc.LastUpdate
                    data = {
                        'id': plc.id,
                        'device_id': plc.device_id,
                        'name': plc.name,
                        'setting': plc.setting,
                        'LastUpdate': plc.LastUpdate
                    }
                    yield f"data: {json.dumps(data, ensure_ascii=False)}\n\n"
            except PLC.DoesNotExist:
                pass
            
            time.sleep(1)
    
    response = StreamingHttpResponse(event_stream(), content_type='text/event-stream')
    response['Cache-Control'] = 'no-cache'
    response['X-Accel-Buffering'] = 'no'
    return response

@csrf_exempt
def toggle_chart_excluded_key(request):
    """Toggle a key in chart excluded keys"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    try:
        key = request.POST.get('key')
        if not key:
            return JsonResponse({'status': 'error', 'message': 'کلید الزامی است'})
        
        excluded = ChartExcludedKeys.objects.filter(key=key).first()
        if excluded:
            excluded.delete()
            return JsonResponse({'status': 'ok', 'action': 'removed', 'message': f'کلید {key} از لیست حذف شد'})
        else:
            ChartExcludedKeys.objects.create(key=key)
            return JsonResponse({'status': 'ok', 'action': 'added', 'message': f'کلید {key} به لیست اضافه شد'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@csrf_exempt
def update_key_order(request):
    """Update order_index for a PLC key"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    try:
        key = request.POST.get('key')
        order_index = request.POST.get('order_index')
        
        if not key:
            return JsonResponse({'status': 'error', 'message': 'کلید الزامی است'})
        
        try:
            order_index = int(order_index) if order_index else 0
        except ValueError:
            order_index = 0
        
        plc_key = PLC_Keys.objects.filter(key=key).first()
        if plc_key:
            plc_key.order_index = order_index
            plc_key.save()
            return JsonResponse({'status': 'ok', 'message': 'ترتیب بروزرسانی شد'})
        else:
            return JsonResponse({'status': 'error', 'message': 'کلید یافت نشد'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@csrf_exempt
def update_keys_order_bulk(request):
    """Bulk update order_index for multiple PLC keys (for drag & drop)"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    try:
        orders_json = request.POST.get('orders')
        if not orders_json:
            return JsonResponse({'status': 'error', 'message': 'داده‌ای ارسال نشده'})
        
        orders = json.loads(orders_json)
        
        with transaction.atomic():
            for key, order_index in orders.items():
                PLC_Keys.objects.filter(key=key).update(order_index=order_index)
        
        return JsonResponse({'status': 'ok', 'message': 'ترتیب بروزرسانی شد'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


def export_all_logs_xlsx(request):
    """Export PLC_Logs with is_running=True to xlsx, filtered by PLC and time range"""
    interval = int(request.GET.get('interval', 1))  # Default 1 second grouping
    plc_id = request.GET.get('plc')
    time_range = request.GET.get('range', '1h')
    from_date = request.GET.get('from_date')  # Jalali date: YYYY/MM/DD
    to_date = request.GET.get('to_date')      # Jalali date: YYYY/MM/DD
    
    # Determine time range
    now = timezone.now()
    
    if from_date and to_date:
        # Custom date range - convert Jalali to Unix timestamp
        try:
            # Parse Jalali dates
            from_parts = from_date.split('/')
            to_parts = to_date.split('/')
            
            # Start of day for from_date
            jdt_from = jdatetime.datetime(
                int(from_parts[0]), int(from_parts[1]), int(from_parts[2]),
                0, 0, 0
            )
            # End of day for to_date
            jdt_to = jdatetime.datetime(
                int(to_parts[0]), int(to_parts[1]), int(to_parts[2]),
                23, 59, 59
            )
            
            # Convert to Gregorian datetime then to timestamp
            threshold_start = jdt_from.togregorian().astimezone(timezone.utc).timestamp()
            threshold_end = jdt_to.togregorian().astimezone(timezone.utc).timestamp()
        except (ValueError, IndexError):
            return HttpResponse('فرمت تاریخ نامعتبر است', content_type='text/plain; charset=utf-8')
    else:
        # Preset time ranges
        time_thresholds = {
            '1h': now - 3600,
            '8h': now - 28800,
            '24h': now - 86400,
            'week': now - 604800,
            'month': now - 2592000,
            'year': now - 31536000
        }
        threshold_start = time_thresholds.get(time_range, now - 3600)
        threshold_end = now
    
    # Build query
    logs_query = PLC_Logs.objects.filter(
        is_running=True,
        CreationDateTime__gte=threshold_start,
        CreationDateTime__lte=threshold_end
    )
    if plc_id:
        logs_query = logs_query.filter(plc_id=plc_id)
    logs = logs_query.select_related('plc', 'roll').order_by('CreationDateTime')
    
    if not logs.exists():
        return HttpResponse('داده‌ای برای خروجی وجود ندارد', content_type='text/plain; charset=utf-8')
    
    # Get key translations from PLC_Keys (key -> fa_name)
    key_map = {}
    for pk in PLC_Keys.objects.all():
        key_map[pk.key] = pk.fa_name or pk.name or pk.key
    
    # Collect all unique keys from json_data
    all_keys = set()
    for log in logs:
        if log.json_data:
            all_keys.update(log.json_data.keys())
    all_keys = sorted(list(all_keys))
    
    # Group logs by roll and time interval, merge json_data
    grouped_data = {}
    
    for log in logs:
        if not log.CreationDateTime:
            continue
        
        roll_id = log.roll_id if log.roll else 0
        interval_ts = int(log.CreationDateTime.timestamp() // interval) * interval
        group_key = (roll_id, interval_ts)
        
        if group_key not in grouped_data:
            plc_name = 'نامشخص'
            if log.plc:
                plc_name = log.plc.name or log.plc.device_id or 'نامشخص'
            grouped_data[group_key] = {
                'plc_name': plc_name,
                'roll_number': log.roll.roll_number if log.roll else '-',
                'timestamp': log.CreationDateTime,
                'data': {}
            }
        
        # Merge json_data (last value wins for each key)
        if log.json_data:
            grouped_data[group_key]['data'].update(log.json_data)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "PLC Logs"
    ws.sheet_view.rightToLeft = True
    
    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Headers: ردیف، دستگاه، شماره رول، زمان، [all keys with fa_name]
    headers = ['ردیف', 'دستگاه', 'شماره رول', 'زمان']
    headers.extend([key_map.get(k, k) for k in all_keys])
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    
    # Data rows (sorted by timestamp)
    sorted_groups = sorted(grouped_data.items(), key=lambda x: x[1]['timestamp'])
    
    for row_idx, (group_key, group_info) in enumerate(sorted_groups, 2):
        # ردیف
        ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
        # دستگاه
        ws.cell(row=row_idx, column=2, value=group_info['plc_name']).border = thin_border
        # شماره رول
        ws.cell(row=row_idx, column=3, value=group_info['roll_number']).border = thin_border
        # زمان (Jalali)
        ts = group_info['timestamp']
        dt_str = jdatetime.datetime.fromtimestamp(ts, timezone.utc).strftime('%Y/%m/%d %H:%M:%S') if ts else ''
        ws.cell(row=row_idx, column=4, value=dt_str).border = thin_border
        
        # Key values
        for col_idx, key in enumerate(all_keys, 5):
            value = group_info['data'].get(key, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response with Jalali filename
    jalali_now = jdatetime.datetime.now(timezone.utc)
    filename = f"plc_logs_{jalali_now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response